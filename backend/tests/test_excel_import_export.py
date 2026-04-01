"""
Test Excel Import/Export Feature for Admin Data Management
Tests: GET /api/admin/excel/sources, GET /api/admin/excel/export/{source}, 
       GET /api/admin/excel/template/{source}, POST /api/admin/excel/import/{source}
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Admin credentials
ADMIN_EMAIL = "admin@musclegrid.in"
ADMIN_PASSWORD = "Muscle@846"

# Expected data sources
EXPECTED_DATA_SOURCES = ["customers", "dealers", "orders", "warranties", "master_skus", "inventory"]


@pytest.fixture(scope="module")
def admin_token():
    """Get admin authentication token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": ADMIN_EMAIL,
        "password": ADMIN_PASSWORD
    })
    assert response.status_code == 200, f"Admin login failed: {response.text}"
    return response.json()["access_token"]


@pytest.fixture(scope="module")
def admin_headers(admin_token):
    """Get headers with admin auth token"""
    return {
        "Authorization": f"Bearer {admin_token}",
        "Content-Type": "application/json"
    }


class TestExcelDataSources:
    """Test GET /api/admin/excel/sources endpoint"""
    
    def test_get_sources_returns_200(self, admin_headers):
        """Test that sources endpoint returns 200"""
        response = requests.get(f"{BASE_URL}/api/admin/excel/sources", headers=admin_headers)
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
    
    def test_get_sources_returns_6_sources(self, admin_headers):
        """Test that exactly 6 data sources are returned"""
        response = requests.get(f"{BASE_URL}/api/admin/excel/sources", headers=admin_headers)
        data = response.json()
        assert len(data) == 6, f"Expected 6 sources, got {len(data)}: {[s['key'] for s in data]}"
    
    def test_get_sources_has_correct_keys(self, admin_headers):
        """Test that all expected data sources are present"""
        response = requests.get(f"{BASE_URL}/api/admin/excel/sources", headers=admin_headers)
        data = response.json()
        source_keys = [s["key"] for s in data]
        
        for expected_key in EXPECTED_DATA_SOURCES:
            assert expected_key in source_keys, f"Missing data source: {expected_key}"
    
    def test_get_sources_has_record_counts(self, admin_headers):
        """Test that each source has record_count field"""
        response = requests.get(f"{BASE_URL}/api/admin/excel/sources", headers=admin_headers)
        data = response.json()
        
        for source in data:
            assert "record_count" in source, f"Missing record_count for {source['key']}"
            assert isinstance(source["record_count"], int), f"record_count should be int for {source['key']}"
    
    def test_get_sources_has_required_fields(self, admin_headers):
        """Test that each source has fields and required_fields"""
        response = requests.get(f"{BASE_URL}/api/admin/excel/sources", headers=admin_headers)
        data = response.json()
        
        for source in data:
            assert "fields" in source, f"Missing fields for {source['key']}"
            assert "required_fields" in source, f"Missing required_fields for {source['key']}"
            assert isinstance(source["fields"], list), f"fields should be list for {source['key']}"
            assert len(source["fields"]) > 0, f"fields should not be empty for {source['key']}"
    
    def test_get_sources_requires_auth(self):
        """Test that sources endpoint requires authentication"""
        response = requests.get(f"{BASE_URL}/api/admin/excel/sources")
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got {response.status_code}"


class TestExcelExport:
    """Test GET /api/admin/excel/export/{data_source} endpoint"""
    
    @pytest.mark.parametrize("source", EXPECTED_DATA_SOURCES)
    def test_export_returns_excel_file(self, admin_headers, source):
        """Test that export returns an Excel file for each data source"""
        response = requests.get(
            f"{BASE_URL}/api/admin/excel/export/{source}", 
            headers=admin_headers
        )
        assert response.status_code == 200, f"Export {source} failed: {response.status_code} - {response.text}"
        
        # Check content type is Excel
        content_type = response.headers.get("content-type", "")
        assert "spreadsheet" in content_type or "excel" in content_type.lower() or "octet-stream" in content_type, \
            f"Expected Excel content type for {source}, got: {content_type}"
        
        # Check content disposition has filename
        content_disp = response.headers.get("content-disposition", "")
        assert "attachment" in content_disp, f"Expected attachment disposition for {source}"
        assert ".xlsx" in content_disp, f"Expected .xlsx filename for {source}"
    
    def test_export_invalid_source_returns_400(self, admin_headers):
        """Test that invalid data source returns 400"""
        response = requests.get(
            f"{BASE_URL}/api/admin/excel/export/invalid_source", 
            headers=admin_headers
        )
        assert response.status_code == 400, f"Expected 400 for invalid source, got {response.status_code}"
    
    def test_export_requires_auth(self):
        """Test that export requires authentication"""
        response = requests.get(f"{BASE_URL}/api/admin/excel/export/customers")
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got {response.status_code}"


class TestExcelTemplate:
    """Test GET /api/admin/excel/template/{data_source} endpoint"""
    
    @pytest.mark.parametrize("source", EXPECTED_DATA_SOURCES)
    def test_template_returns_excel_file(self, admin_headers, source):
        """Test that template returns an Excel file for each data source"""
        response = requests.get(
            f"{BASE_URL}/api/admin/excel/template/{source}", 
            headers=admin_headers
        )
        assert response.status_code == 200, f"Template {source} failed: {response.status_code} - {response.text}"
        
        # Check content type is Excel
        content_type = response.headers.get("content-type", "")
        assert "spreadsheet" in content_type or "excel" in content_type.lower() or "octet-stream" in content_type, \
            f"Expected Excel content type for {source}, got: {content_type}"
        
        # Check content disposition has filename
        content_disp = response.headers.get("content-disposition", "")
        assert "attachment" in content_disp, f"Expected attachment disposition for {source}"
        assert "_template.xlsx" in content_disp, f"Expected _template.xlsx filename for {source}"
    
    def test_template_invalid_source_returns_400(self, admin_headers):
        """Test that invalid data source returns 400"""
        response = requests.get(
            f"{BASE_URL}/api/admin/excel/template/invalid_source", 
            headers=admin_headers
        )
        assert response.status_code == 400, f"Expected 400 for invalid source, got {response.status_code}"
    
    def test_template_requires_auth(self):
        """Test that template requires authentication"""
        response = requests.get(f"{BASE_URL}/api/admin/excel/template/master_skus")
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got {response.status_code}"


class TestExcelImport:
    """Test POST /api/admin/excel/import/{data_source} endpoint"""
    
    def test_import_requires_auth(self):
        """Test that import requires authentication"""
        response = requests.post(f"{BASE_URL}/api/admin/excel/import/master_skus")
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got {response.status_code}"
    
    def test_import_invalid_source_returns_400(self, admin_headers):
        """Test that invalid data source returns 400"""
        # Create a minimal Excel file
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["test"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"mode": "merge"}
        
        headers = {"Authorization": admin_headers["Authorization"]}
        response = requests.post(
            f"{BASE_URL}/api/admin/excel/import/invalid_source",
            headers=headers,
            files=files,
            data=data
        )
        assert response.status_code == 400, f"Expected 400 for invalid source, got {response.status_code}"
    
    def test_import_non_excel_file_returns_400(self, admin_headers):
        """Test that non-Excel file returns 400"""
        files = {"file": ("test.txt", io.BytesIO(b"not an excel file"), "text/plain")}
        data = {"mode": "merge"}
        
        headers = {"Authorization": admin_headers["Authorization"]}
        response = requests.post(
            f"{BASE_URL}/api/admin/excel/import/master_skus",
            headers=headers,
            files=files,
            data=data
        )
        assert response.status_code == 400, f"Expected 400 for non-Excel file, got {response.status_code}"
    
    def test_import_master_skus_merge_mode(self, admin_headers):
        """Test importing master_skus in merge mode with valid data"""
        import openpyxl
        
        # Create Excel file with valid master_sku data
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers matching the expected fields
        headers_row = ["id", "sku_code", "name", "description", "category", "subcategory", 
                       "hsn_code", "mrp", "dealer_price", "unit", "min_stock", "is_active", "created_at"]
        ws.append(headers_row)
        
        # Add a test row with unique SKU code
        import uuid
        test_sku_code = f"TEST-SKU-{uuid.uuid4().hex[:8].upper()}"
        ws.append([
            "",  # id - leave empty for new
            test_sku_code,
            "Test Product for Import",
            "Test description",
            "Inverter",
            "",
            "8504",
            "10000",
            "8000",
            "pcs",
            "5",
            "True",
            ""  # created_at - auto-filled
        ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test_import.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"mode": "merge"}
        
        auth_headers = {"Authorization": admin_headers["Authorization"]}
        response = requests.post(
            f"{BASE_URL}/api/admin/excel/import/master_skus",
            headers=auth_headers,
            files=files,
            data=data
        )
        
        assert response.status_code == 200, f"Import failed: {response.status_code} - {response.text}"
        
        result = response.json()
        assert result["success"] == True, f"Import not successful: {result}"
        assert result["data_source"] == "master_skus"
        assert result["mode"] == "merge"
        assert "imported" in result
        assert "updated" in result
        assert "total_processed" in result
        
        # Verify at least 1 record was processed
        assert result["total_processed"] >= 1, f"Expected at least 1 record processed, got {result['total_processed']}"
    
    def test_import_missing_required_fields_returns_400(self, admin_headers):
        """Test that import with missing required fields returns 400"""
        import openpyxl
        
        # Create Excel file with missing required fields
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers
        ws.append(["id", "sku_code", "name", "category"])
        
        # Row with missing required field (name is empty)
        ws.append(["", "SKU-MISSING", "", "Inverter"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test_missing.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"mode": "merge"}
        
        auth_headers = {"Authorization": admin_headers["Authorization"]}
        response = requests.post(
            f"{BASE_URL}/api/admin/excel/import/master_skus",
            headers=auth_headers,
            files=files,
            data=data
        )
        
        # Should return 400 due to missing required field
        assert response.status_code == 400, f"Expected 400 for missing required fields, got {response.status_code}: {response.text}"


class TestExcelImportCustomers:
    """Test Excel import for customers data source"""
    
    def test_import_customers_merge_mode(self, admin_headers):
        """Test importing customers in merge mode"""
        import openpyxl
        import uuid
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers for customers
        headers_row = ["id", "email", "first_name", "last_name", "phone", "address", "city", "state", "pincode", "created_at"]
        ws.append(headers_row)
        
        # Add test customer with unique email
        test_email = f"test.import.{uuid.uuid4().hex[:8]}@example.com"
        ws.append([
            "",  # id
            test_email,
            "Test",
            "Import",
            "9876543210",
            "123 Test Street",
            "Mumbai",
            "Maharashtra",
            "400001",
            ""  # created_at
        ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("customers_import.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"mode": "merge"}
        
        auth_headers = {"Authorization": admin_headers["Authorization"]}
        response = requests.post(
            f"{BASE_URL}/api/admin/excel/import/customers",
            headers=auth_headers,
            files=files,
            data=data
        )
        
        assert response.status_code == 200, f"Customer import failed: {response.status_code} - {response.text}"
        
        result = response.json()
        assert result["success"] == True
        assert result["data_source"] == "customers"


class TestExcelRoundTrip:
    """Test export -> modify -> import workflow"""
    
    def test_export_and_reimport_master_skus(self, admin_headers):
        """Test that exported data can be re-imported"""
        # Step 1: Export master_skus
        export_response = requests.get(
            f"{BASE_URL}/api/admin/excel/export/master_skus",
            headers=admin_headers
        )
        assert export_response.status_code == 200, f"Export failed: {export_response.status_code}"
        
        # Step 2: Re-import the same file (should update existing records)
        files = {"file": ("reimport.xlsx", io.BytesIO(export_response.content), 
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"mode": "merge"}
        
        auth_headers = {"Authorization": admin_headers["Authorization"]}
        import_response = requests.post(
            f"{BASE_URL}/api/admin/excel/import/master_skus",
            headers=auth_headers,
            files=files,
            data=data
        )
        
        # Should succeed (may have 0 imported if all records already exist)
        assert import_response.status_code == 200, f"Re-import failed: {import_response.status_code} - {import_response.text}"
        
        result = import_response.json()
        assert result["success"] == True
        # In merge mode, existing records should be updated
        assert result["total_processed"] >= 0


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
