"""
Excel Export/Import for Bulk CRM Data Updates
Allows exporting orders with missing data and importing filled data back
"""

import io
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_missing_data_excel(orders: List[Dict[str, Any]]) -> bytes:
    """
    Create an Excel file with orders that have missing data.
    Columns are designed for easy filling by AI or manual entry.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders Missing Data"
    
    # Define columns
    columns = [
        ("A", "order_id", "Order ID", 20),
        ("B", "amazon_order_id", "Amazon Order ID", 25),
        ("C", "status", "Current Status", 15),
        ("D", "customer_name", "Customer Name", 25),
        ("E", "customer_phone", "Phone (10 digits)", 15),
        ("F", "address", "Address", 40),
        ("G", "city", "City", 15),
        ("H", "state", "State", 15),
        ("I", "pincode", "Pincode (6 digits)", 12),
        ("J", "tracking_id", "Tracking ID (AWB)", 20),
        ("K", "carrier", "Carrier (Delhivery/BlueDart)", 20),
        ("L", "product_name", "Product Name", 35),
        ("M", "sku_code", "SKU Code", 15),
        ("N", "quantity", "Qty", 5),
        ("O", "order_total", "Order Total (₹)", 12),
        ("P", "firm_id", "Firm ID (DO NOT EDIT)", 20),
        ("Q", "pf_id", "PF ID (DO NOT EDIT)", 25),
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    readonly_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    missing_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Yellow for missing
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_letter, field_name, header_text, width in columns:
        cell = ws[f"{col_letter}1"]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[col_letter].width = width
    
    # Write data rows
    for row_idx, order in enumerate(orders, start=2):
        for col_letter, field_name, _, _ in columns:
            cell = ws[f"{col_letter}{row_idx}"]
            value = order.get(field_name, "")
            cell.value = value if value else ""
            cell.border = thin_border
            
            # Highlight missing fields in yellow
            if field_name in ["customer_name", "customer_phone", "address", "city", "state", "pincode", "tracking_id"]:
                if not value:
                    cell.fill = missing_fill
            
            # Mark readonly columns
            if field_name in ["firm_id", "pf_id", "order_id"]:
                cell.fill = readonly_fill
    
    # Add instructions sheet
    instructions = wb.create_sheet("Instructions")
    instructions["A1"] = "BULK DATA UPDATE INSTRUCTIONS"
    instructions["A1"].font = Font(bold=True, size=14)
    
    instructions["A3"] = "1. Fill in the YELLOW highlighted cells with missing data"
    instructions["A4"] = "2. DO NOT modify grey columns (Order ID, Firm ID, PF ID)"
    instructions["A5"] = "3. Phone must be 10 digits (e.g., 9876543210)"
    instructions["A6"] = "4. Pincode must be 6 digits (e.g., 110001)"
    instructions["A7"] = "5. State should be full name (e.g., Maharashtra, Karnataka)"
    instructions["A8"] = "6. Tracking ID = AWB number from courier"
    instructions["A9"] = "7. Carrier = Delhivery, BlueDart, DTDC, Ecom Express, etc."
    instructions["A10"] = ""
    instructions["A11"] = "IMPORTANT:"
    instructions["A11"].font = Font(bold=True, color="DC2626")
    instructions["A12"] = "- Orders with TRACKING ID will be auto-marked as DISPATCHED"
    instructions["A13"] = "- Invoice and Label PDFs are OPTIONAL (can upload later)"
    instructions["A14"] = "- Save file as .xlsx before uploading back"
    
    instructions.column_dimensions["A"].width = 60
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_import_excel(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Parse an uploaded Excel file and extract order data for update.
    Returns list of orders with their data.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    
    # Find the data sheet (first sheet or "Orders Missing Data")
    ws = None
    if "Orders Missing Data" in wb.sheetnames:
        ws = wb["Orders Missing Data"]
    else:
        ws = wb.active
    
    # Map column letters to field names
    column_map = {
        "A": "order_id",
        "B": "amazon_order_id",
        "C": "status",
        "D": "customer_name",
        "E": "customer_phone",
        "F": "address",
        "G": "city",
        "H": "state",
        "I": "pincode",
        "J": "tracking_id",
        "K": "carrier",
        "L": "product_name",
        "M": "sku_code",
        "N": "quantity",
        "O": "order_total",
        "P": "firm_id",
        "Q": "pf_id",
    }
    
    orders = []
    
    # Skip header row, read data rows
    for row_idx in range(2, ws.max_row + 1):
        order = {}
        has_data = False
        
        for col_letter, field_name in column_map.items():
            cell_value = ws[f"{col_letter}{row_idx}"].value
            if cell_value is not None:
                # Convert to string and strip whitespace
                if isinstance(cell_value, (int, float)):
                    cell_value = str(int(cell_value)) if field_name in ["customer_phone", "pincode", "quantity"] else str(cell_value)
                else:
                    cell_value = str(cell_value).strip()
                
                if cell_value:
                    order[field_name] = cell_value
                    has_data = True
        
        # Only add rows that have at least order_id or pf_id
        if has_data and (order.get("order_id") or order.get("pf_id") or order.get("amazon_order_id")):
            orders.append(order)
    
    return orders


def validate_import_data(orders: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Validate imported data and return validation results.
    """
    errors = []
    warnings = []
    valid_orders = []
    
    for idx, order in enumerate(orders):
        row_num = idx + 2  # Excel row number (1-indexed + header)
        row_errors = []
        row_warnings = []
        
        # Must have identifier
        if not (order.get("order_id") or order.get("pf_id") or order.get("amazon_order_id")):
            row_errors.append("Missing order identifier (Order ID, PF ID, or Amazon Order ID)")
        
        # Validate phone if provided
        phone = order.get("customer_phone", "")
        if phone:
            phone_clean = ''.join(filter(str.isdigit, phone))
            if len(phone_clean) != 10:
                row_warnings.append(f"Phone '{phone}' should be 10 digits")
            else:
                order["customer_phone"] = phone_clean
        
        # Validate pincode if provided
        pincode = order.get("pincode", "")
        if pincode:
            pincode_clean = ''.join(filter(str.isdigit, pincode))
            if len(pincode_clean) != 6:
                row_warnings.append(f"Pincode '{pincode}' should be 6 digits")
            else:
                order["pincode"] = pincode_clean
        
        # Check if has tracking ID (will be auto-dispatched)
        if order.get("tracking_id"):
            order["auto_dispatch"] = True
        
        if row_errors:
            errors.append({"row": row_num, "errors": row_errors})
        else:
            if row_warnings:
                warnings.append({"row": row_num, "warnings": row_warnings})
            valid_orders.append(order)
    
    return {
        "valid_count": len(valid_orders),
        "error_count": len(errors),
        "warning_count": len(warnings),
        "valid_orders": valid_orders,
        "errors": errors,
        "warnings": warnings
    }
